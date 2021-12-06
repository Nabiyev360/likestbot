from openpyxl import load_workbook

# STANDART FUNKSIYALAR
# Excel database'ga ulanish
def dbxl_connect(filename):
    wb = load_workbook(filename)
    sheet = wb.active
    return sheet


# Ma'lumotni olish, cordinate orqali
def get_data(cordinate):
    sheet = dbxl_connect('data/main.xlsx')
    data = sheet[cordinate].value
    return data


# Ma'lumot qo'shish uchun, cordinata va ma'lumot kiritish orqali
def add_data(cordinate, item):
    wb = load_workbook('data/main.xlsx')
    sheet = wb.active
    sheet[cordinate] = item
    wb.save('data/main.xlsx')


def add_photo(file_id, unique_id, lang):
    wb = load_workbook('data/main.xlsx') 
    sheet = wb.active
    
    IDs = []
    for row in sheet.rows:
        IDs.append(str(row[1].value))
    for id in IDs:
        N = len(IDs) + 1
        # print(id  )
    # print(IDs)
    
    sheet[f'A{N}'] = lang
    sheet[f'B{N}'] = file_id
    sheet[f'C{N}'] = unique_id
    sheet[f'D{N}'] = 0
    # sheet[f'G{N}'] = []

    wb.save('data/main.xlsx')


#   QO'SHIMCHA FUNKSIYALAR
# Foydalanuvchi qo'shish uchun. id, fullname kiritish orqali


def likeup_photo(unique_id, user_id):
    wb = load_workbook('data/main.xlsx') 
    sheet = wb.active
    IDs = []
    for row in sheet.rows:
        if row[2].value != None:
            IDs.append(row[2].value)
    
    if unique_id in IDs:
        N = IDs.index(unique_id) + 1
        
        
        # Avval bosgan yoki bosmaganlikni aniqlash
        if sheet[f'E{N}'].value == None:
            sheet[f'E{N}'].value = f"{user_id}, "
            sheet[f'D{N}'].value += 1
        
        elif str(user_id) not in sheet[f'E{N}'].value:
            sheet[f'E{N}'].value += f"{user_id}, "
            sheet[f'D{N}'].value += 1
        else:
            pass
         
        wb.save('data/main.xlsx')    
        
        return sheet[f'D{N}'].value
    else:
        return None

def get_like(unique_id):
    sheet = dbxl_connect('data/main.xlsx')
    IDs = []
    for row in sheet.rows:
        if row[2].value != None:
            IDs.append(row[2].value)
    
    if unique_id in IDs:
        N = IDs.index(unique_id) + 1
        likes = sheet[f'D{N}'].value 
        
        return likes
        
        
def get_photo(unique_id):
    sheet = dbxl_connect('data/main.xlsx')
    IDs = []
    for row in sheet.rows:
        if row[2].value != None:
            IDs.append(row[2].value)
    
    if unique_id in IDs:
        N = IDs.index(unique_id) + 1
        photo_id = sheet[f'B{N}'].value
        
        return photo_id
        
        
def get_caption(unique_id):
    sheet = dbxl_connect('data/main.xlsx')
    IDs = []
    for row in sheet.rows:
        if row[2].value != None:
            IDs.append(row[2].value)
    
    if unique_id in IDs:
        N = IDs.index(unique_id) + 1
        caption = sheet[f'F{N}'].value
        
        return caption


def set_language(lang):
    add_data(cordinate="A1", item=lang)

    
def get_language(unique_id):
    sheet = dbxl_connect('data/main.xlsx')
    IDs = []
    for row in sheet.rows:
        if row[2].value != None:
            IDs.append(row[2].value)
    
    if unique_id in IDs:
        N = IDs.index(unique_id) + 1
        lang = sheet[f'A{N}'].value
        
        return lang